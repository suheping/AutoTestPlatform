#-*- coding:utf-8 -*-
# author:peace
# datetime:2018/9/29 17:53
# file:copyXls
# desc: 复制xls，填写测试结果工具类

import openpyxl
def copyXls(xlsPath1,xlsPath2):
    '''复制xls，复制xlsPath1保存为xlsPath2'''
    wb2 = openpyxl.Workbook()
    wb2.save(xlsPath2)
    # 打开测试数据文件、测试结果文件
    wb1 = openpyxl.load_workbook(xlsPath1,read_only=True)
    wb2 = openpyxl.load_workbook(xlsPath2)

    # 取到所有sheet的名字
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames

    # 读取第一个sheet
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]

    # 测试数据文件中的行数、列数
    rowCount = sheet1.max_row
    colCount = sheet1.max_column

    for m in list(range(1, rowCount + 1)):
        for n in list(range(97, 97 + colCount)):  # chr(97)='a'
            n = chr(n)  # ASCII字符
            i = '%s%d' % (n, m)  # 单元格编号
            cell1 = sheet1[i].value  # 获取data单元格数据
            sheet2[i].value = cell1  # 赋值到test单元格

    wb2.save(xlsPath2)  # 保存数据
    wb1.close()  # 关闭excel
    wb2.close()

class writeXls(object):
    '''修改excel数据'''

    def __init__(self, filename):
        self.filename = filename
        self.wb = openpyxl.load_workbook(self.filename)
        self.ws = self.wb.active  # 激活sheet

    def write(self, row_n, col_n, value):
        '''写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"'''
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)

if __name__ == '__main__':
    copyXls('../data/case1.xlsx','../report/case1_result.xlsx')
    wt = writeXls('../report/case1_result.xlsx')
    wt.write(3,1,'shp')